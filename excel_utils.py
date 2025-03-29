import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
import db_utils
from psycopg2.extras import DictCursor

def download_weekly_schedule_excel(start_date, end_date):
    """
    週間予定をExcelファイルとしてダウンロードする
    """
    try:
        conn = db_utils.get_db_connection()
        cur = conn.cursor(cursor_factory=DictCursor)

        # 週間予定データを取得
        cur.execute("""
            SELECT 投稿者, 開始日, 終了日, 月曜日, 火曜日, 水曜日, 木曜日, 金曜日, 土曜日, 日曜日, 投稿日時
            FROM weekly_schedules
            WHERE 開始日 = %s AND 終了日 = %s
        """, (start_date, end_date))
        rows = cur.fetchall()

        if not rows:
            return None  # データがない場合は None を返す

        wb = openpyxl.Workbook()
        ws = wb.active

        # ヘッダー行
        headers = ["投稿者", "開始日", "終了日", "月曜日", "火曜日", "水曜日", "木曜日", "金曜日", "土曜日", "日曜日", "投稿日時"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # データ行
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # ファイルをバイト列として保存
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    except Exception as e:
        print(f"エラーが発生しました: {e}")
        return None

    finally:
        if 'cur' in locals() and cur is not None:
            cur.close()
        if 'conn' in locals() and conn is not None:
            conn.close()
